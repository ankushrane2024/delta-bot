import os
import json
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

USD_INR_RATE = 83.0  # Fixed exchange rate as default

def generate_pdf_report(data, filepath):
    """Generates a premium, beautiful PDF report for daily trading."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom styles for a premium look
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0f172a'), # Slate 900
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1e40af'), # Blue 800
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#334155') # Slate 700
    )
    
    body_bold = ParagraphStyle(
        'BodyBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )

    # Document Header
    story.append(Paragraph(f"Daily Trade Report — {data['date']}", title_style))
    story.append(Paragraph(f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} IST", body_style))
    story.append(Spacer(1, 15))
    
    # Summary Table
    story.append(Paragraph("1. Executive Summary", section_style))
    summary = data['summary']
    summary_data = [
        [Paragraph("Metric", body_bold), Paragraph("Value", body_bold), Paragraph("Metric", body_bold), Paragraph("Value", body_bold)],
        ["Total Trades", str(summary['total_trades']), "Win Rate", f"{summary['win_rate']:.2f}%"],
        ["Net P&L (USD)", f"${summary['net_pnl_usd']:.2f}", "Net P&L (INR)", f"Rs. {summary['net_pnl_inr']:,.2f}"],
        ["Max Drawdown", f"{summary['max_drawdown']:.2f}%", "Market Regime", summary['market_regime']],
        ["Filter Status (ADX)", "ON" if summary['regime_filter_enabled'] else "OFF", "", ""]
    ]
    
    t_summary = Table(summary_data, colWidths=[130, 130, 130, 130])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 15))
    
    # Trade Details
    story.append(Paragraph("2. Trade Details", section_style))
    trades = data['trades']
    if not trades:
        story.append(Paragraph("No trades executed today.", body_style))
    else:
        trades_headers = ["Entry", "Exit", "Call Leg", "Put Leg", "Tot Prm", "P&L", "Reason"]
        trades_rows = [[Paragraph(h, body_bold) for h in trades_headers]]
        for t in trades:
            # Shorten strikes and add entry/exit prices for PDF fit
            call_s = t['call_strike'].replace('C-BTC-', 'C-')
            put_s = t['put_strike'].replace('P-BTC-', 'P-')
            
            call_text = f"{call_s}<br/>${t.get('call_entry_price', 0):.2f} → ${t.get('call_exit_price', 0):.2f}"
            put_text = f"{put_s}<br/>${t.get('put_entry_price', 0):.2f} → ${t.get('put_exit_price', 0):.2f}"
            
            pnl_val = t['pnl_usd']
            pnl_color = '#10b981' if pnl_val >= 0 else '#ef4444'
            pnl_text = f"<font color='{pnl_color}'>${pnl_val:.2f}</font>"
            
            hedge_pnl = t.get('hedge_pnl', 0.0)
            if hedge_pnl != 0.0:
                h_color = '#10b981' if hedge_pnl >= 0 else '#ef4444'
                pnl_text += f"<br/><font size='8'>Hedge: <font color='{h_color}'>${hedge_pnl:.2f}</font></font>"
                
            max_pnl = t.get('max_pnl_pct', 0.0) * 100
            min_pnl = t.get('min_pnl_pct', 0.0) * 100
            if max_pnl != -99900.0 and min_pnl != 99900.0:
                pnl_text += f"<br/><font size='7' color='#64748b'>Peak: +{max_pnl:.1f}%<br/>Trough: {min_pnl:.1f}%</font>"
            
            trades_rows.append([
                t['entry_time'].split('T')[-1][:8] if 'T' in t['entry_time'] else t['entry_time'],
                t['exit_time'].split('T')[-1][:8] if 'T' in t['exit_time'] else t['exit_time'],
                Paragraph(call_text, body_style),
                Paragraph(put_text, body_style),
                f"${t['entry_premium']:.4f}",
                Paragraph(pnl_text, body_style),
                t['exit_reason']
            ])
        t_trades = Table(trades_rows, colWidths=[60, 60, 105, 105, 55, 55, 90])
        t_trades.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ]))
        story.append(t_trades)
    story.append(Spacer(1, 15))
    
    # Risk Metrics & Market Conditions in 2 columns
    story.append(Paragraph("3. Risk Metrics & Market Conditions", section_style))
    risk = data['risk']
    market = data['market']
    
    detail_data = [
        [Paragraph("Risk Metric", body_bold), Paragraph("Value", body_bold), Paragraph("Market Condition", body_bold), Paragraph("Value", body_bold)],
        ["Loss Limit Hit", "Yes" if risk['daily_loss_limit_hit'] else "No", "ADX Value", f"{market['adx']:.2f}"],
        ["SL Hits Today", str(risk['sl_hits']), "IV Level", f"{market['iv']:.4f}"],
        ["Hedging Activity", risk['hedging_activity'], "High Impact News", market['news'] or "None"]
    ]
    t_detail = Table(detail_data, colWidths=[130, 130, 130, 130])
    t_detail.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    story.append(t_detail)
    
    doc.build(story)

def generate_xlsx_report(data, filepath):
    """Generates a detailed Excel report matching the daily data."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    
    # Styling definitions
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1E3A8A")
    section_font = Font(name=font_family, size=12, bold=True, color="1E40AF")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    summary_label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 1. Document Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Daily Trading Report — {data['date']}"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} IST"
    ws["A2"].font = Font(name=font_family, size=9, italic=True, color="64748B")
    ws.row_dimensions[2].height = 18
    
    # 2. Executive Summary
    ws["A4"] = "Executive Summary"
    ws["A4"].font = section_font
    ws.row_dimensions[4].height = 20
    
    summary = data['summary']
    summary_items = [
        ("Total Trades", summary['total_trades']),
        ("Win Rate", f"{summary['win_rate']:.2f}%"),
        ("Net P&L (USD)", summary['net_pnl_usd']),
        ("Net P&L (INR)", summary['net_pnl_inr']),
        ("Max Drawdown", f"{summary['max_drawdown']:.2f}%"),
        ("Market Regime", summary['market_regime']),
        ("Regime Filter Status", "ON" if summary['regime_filter_enabled'] else "OFF")
    ]
    
    row_idx = 5
    for label, val in summary_items:
        ws.cell(row=row_idx, column=1, value=label).font = bold_data_font
        ws.cell(row=row_idx, column=1).fill = summary_label_fill
        ws.cell(row=row_idx, column=1).border = thin_border
        
        c = ws.cell(row=row_idx, column=2, value=val)
        c.font = data_font
        c.border = thin_border
        if isinstance(val, (int, float)):
            c.number_format = "$#,##0.00" if "USD" in label else "#,##0.00"
            if "INR" in label:
                c.number_format = "Rs. #,##0.00"
            c.alignment = align_right
        else:
            c.alignment = align_left
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
        
    # 3. Trade Details
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Trade Details").font = section_font
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1
    
    headers = ["Entry Time", "Exit Time", "Call Strike", "Call Entry $", "Call Exit $", "Put Strike", "Put Entry $", "Put Exit $", "Total Premium", "P&L (USD)", "Hedge PnL ($)", "Max Peak %", "Max Trough %", "Exit Reason"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1
    
    trades = data['trades']
    if not trades:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=11)
        ws.cell(row=row_idx, column=1, value="No trades executed today.").font = data_font
        ws.cell(row=row_idx, column=1).alignment = align_left
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
    else:
        for t in trades:
            row_data = [
                t['entry_time'],
                t['exit_time'],
                t['call_strike'],
                t.get('call_entry_price', 0.0),
                t.get('call_exit_price', 0.0),
                t['put_strike'],
                t.get('put_entry_price', 0.0),
                t.get('put_exit_price', 0.0),
                t['entry_premium'],
                t['pnl_usd'],
                t.get('hedge_pnl', 0.0),
                t.get('max_pnl_pct', 0.0) * 100 if t.get('max_pnl_pct', 0.0) != -999.0 else 0.0,
                t.get('min_pnl_pct', 0.0) * 100 if t.get('min_pnl_pct', 0.0) != 999.0 else 0.0,
                t['exit_reason']
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = align_left
                
                # Format numbers
                if col_idx in [4, 5, 7, 8, 9]:
                    cell.number_format = "$0.00"
                    cell.alignment = align_right
                elif col_idx in [10, 11]:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = align_right
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.font = Font(name=font_family, size=10, bold=True, color="10b981" if cell.value >= 0 else "ef4444")
                elif col_idx in [12, 13]:
                    cell.number_format = "0.00%"
                    cell.alignment = align_right
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.value = cell.value / 100.0  # Excel percentage format multiplies by 100
                        cell.font = Font(name=font_family, size=9, color="10b981" if cell.value >= 0 else "ef4444")
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            
    # 4. Risk & Market Metrics
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Risk & Market Metrics").font = section_font
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1
    
    risk = data['risk']
    market = data['market']
    metrics_data = [
        ("Daily Loss Limit Hit", "Yes" if risk['daily_loss_limit_hit'] else "No"),
        ("SL Hits", risk['sl_hits']),
        ("Hedging Activity", risk['hedging_activity']),
        ("ADX Value", market['adx']),
        ("IV Level", market['iv']),
        ("High Impact News", market['news'] or "None")
    ]
    for label, val in metrics_data:
        ws.cell(row=row_idx, column=1, value=label).font = bold_data_font
        ws.cell(row=row_idx, column=1).fill = summary_label_fill
        ws.cell(row=row_idx, column=1).border = thin_border
        
        c = ws.cell(row=row_idx, column=2, value=val)
        c.font = data_font
        c.border = thin_border
        c.alignment = align_left
        if isinstance(val, (int, float)):
            c.alignment = align_right
            c.number_format = "0.0000" if "IV" in label else "0.00"
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
        
    # Auto-fit column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(filepath)

def save_report_data(data):
    """Saves daily report structure to daily_reports.json."""
    filepath = 'daily_reports.json'
    history = {}
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                history = json.load(f)
        except Exception:
            pass
            
    history[data['date']] = data
    
    # Sort history so newest reports are listed first
    sorted_history = dict(sorted(history.items(), key=lambda x: x[0], reverse=True))
    
    with open(filepath, 'w') as f:
        json.dump(sorted_history, f, indent=4)
        
    import db_manager
    db_manager.trigger_cloud_sync()
